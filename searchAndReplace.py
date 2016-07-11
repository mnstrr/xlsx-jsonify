import openpyxl
import json
import ftfy
import codecs
import warnings


class SearchAndReplace:
    def __init__(self, xlsx_doc, json_doc, json_doc_new, sheet_name, startrow, searchcolumn, replacecolumn, blacklist):
        self.__xlsx_doc = xlsx_doc
        self.__json_doc = json_doc
        self.__json_doc_new = json_doc_new
        self.__sheet_name = sheet_name
        self.__startrow = startrow
        self.__searchcolumn = searchcolumn
        self.__replacecolumn = replacecolumn
        self.__blacklist = blacklist
        self.__counter = 0
        self.__data = ""
        self.__newdata = ""

        self.__main()

    def __main(self):

        json_data = open(self.__json_doc, encoding="utf8")

        self.__data = json.load(json_data)
        self.__newdata = self.__data

        print('Starting search and replace. Hold your horses.')
        self.__crawlexcel()

        output_file = codecs.open(self.__json_doc_new, "w", encoding="utf8")
        json.dump(self.__newdata, output_file, indent=2, ensure_ascii=False)

        print(' New jsonfile \"' + self.__json_doc_new + '" was created. ' + str(
            self.__counter) + ' substitutions in total.')

    def __crawlexcel(self):

        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(self.__xlsx_doc)
        sheet = wb.get_sheet_by_name(self.__sheet_name)

        for rowNum in range(self.__searchcolumn, sheet.max_row):  # skip the first row
            progress = int(round((rowNum / (sheet.max_row - 2)) * 100))
            self.__update_progress(progress)
            text = sheet.cell(row=rowNum, column=self.__searchcolumn).value
            translation = sheet.cell(row=rowNum, column=self.__replacecolumn).value
            if text is not None and translation is not None:
                searchstring = ftfy.fix_encoding(text).lower().replace(" ", "").replace("\"", "")
                replacestring = ftfy.fix_encoding(translation)
                self.__newdata = self.__checkdict(self.__newdata, searchstring, replacestring)

    def __checkdict(self, collection, k, v):
        collectiontemp = collection
        for index in collectiontemp.keys():
            if "href" not in index:
                collectiontemp[index] = self.__checkitem(collectiontemp, k, v, index)
        return collectiontemp

    def __checklist(self, collection, k, v):
        collectiontemp = collection
        for index in range(len(collectiontemp)):
            collectiontemp[index] = self.__checkitem(collectiontemp, k, v, index)
        return collectiontemp

    def __checkitem(self, collection, k, v, index):
        collectiontemp = collection
        item = collectiontemp[index]

        if type(item) is str:
            if ftfy.fix_encoding(item).lower().replace(" ", "").replace("\"", "") == k:
                self.__counter += 1
                return ftfy.fix_encoding(v)
            else:
                return ftfy.fix_encoding(item)
        if type(item) is dict:
            collectiontemp = self.__checkdict(item, k, v)
        if type(item) is list:
            collectiontemp = self.__checklist(item, k, v)
        if type(item) is int or type(item) is float:
            return item

        return collectiontemp

    def ___naivereplacement(self, text, translation):
        self.__newdata = self.__newdata.replace(text, translation)

    def __update_progress(self, progress):
        print('\r[{0:10}]{1:>2}%'.format('#' * int(progress * 10 / 100), progress), end='')
